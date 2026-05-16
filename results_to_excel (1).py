"""
results_to_excel.py
Run AFTER test_job_profit_os.py:
    python results_to_excel.py
Reads test_results.json → writes Job_Profit_OS_Test_Results.xlsx
"""

import json, re, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

CLR = {
    "navy":   "1F3864", "blue":  "2F5496", "sky":   "D6E4F0",
    "pass":   "70AD47", "fail":  "C00000", "error": "FF0000",
    "skip":   "ED7D31", "white": "FFFFFF", "grey":  "F2F2F2",
    "yellow": "FFF2CC", "dgrey": "595959", "green": "375623",
}

STATUS_CLR = {
    "PASS":       (CLR["pass"],  CLR["white"]),
    "FAIL":       (CLR["fail"],  CLR["white"]),
    "ERROR":      (CLR["error"], CLR["white"]),
    "SKIP":       (CLR["skip"],  CLR["white"]),
    "NOT TESTED": (CLR["dgrey"], CLR["white"]),
}

def hdr(ws, row, col, value, bg=CLR["navy"], fg="FFFFFF", size=11, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, size=size, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h=h, v="center")
    c.border = thin_border()
    return c

def cell(ws, row, col, value, bg=CLR["white"], bold=False, h="left", wrap=False, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, color=color)
    c.fill = fill(bg)
    c.alignment = align(h=h, v="center", wrap=wrap)
    c.border = thin_border()
    return c

def status_cell(ws, row, col, status):
    bg, fg = STATUS_CLR.get(status, (CLR["dgrey"], CLR["white"]))
    c = ws.cell(row=row, column=col, value=status)
    c.font = Font(name="Arial", bold=True, size=10, color=fg)
    c.fill = fill(bg)
    c.alignment = align(h="center", v="center")
    c.border = thin_border()
    return c

# ── Load results ──────────────────────────────────────────────────────────────
with open("test_results.json") as f:
    results = json.load(f)

total   = len(results)
passed  = sum(1 for r in results if r["status"] == "PASS")
failed  = sum(1 for r in results if r["status"] == "FAIL")
errors  = sum(1 for r in results if r["status"] == "ERROR")
skipped = sum(1 for r in results if r["status"] == "SKIP")
pass_rate = f"{passed/max(total,1)*100:.1f}%"

wb = Workbook()

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1 – SUMMARY
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"

ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "JOB PROFIT OS — AUTOMATED TEST RESULTS"
c.font = Font(name="Arial", bold=True, size=16, color=CLR["white"])
c.fill = fill(CLR["navy"])
c.alignment = align(h="center")
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = f"Run: {datetime.now().strftime('%d %b %Y %H:%M')}  |  URL: https://job-profit-os-git-main-26jzk94vgp-techs-projects.vercel.app"
c.font = font(italic=True, color=CLR["dgrey"], size=10)
c.fill = fill(CLR["sky"])
c.alignment = align(h="center")
ws.row_dimensions[2].height = 18

# Score cards row
score_data = [
    ("Total Tests", total,   CLR["blue"]),
    ("✅ PASS",     passed,  CLR["pass"]),
    ("❌ FAIL",     failed,  CLR["fail"]),
    ("💥 ERROR",   errors,  "C55A11"),
    ("⏭  SKIP",   skipped, CLR["dgrey"]),
    ("Pass Rate",  pass_rate, CLR["navy"]),
]

ws.row_dimensions[3].height = 6  # spacer
for i, (lbl, val, clr) in enumerate(score_data):
    col = i + 1
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 40
    c1 = ws.cell(row=4, column=col, value=lbl)
    c1.font = Font(name="Arial", bold=True, size=10, color=CLR["white"])
    c1.fill = fill(clr)
    c1.alignment = align(h="center")
    c1.border = thin_border()
    c2 = ws.cell(row=5, column=col, value=val)
    c2.font = Font(name="Arial", bold=True, size=20, color=clr)
    c2.fill = fill(CLR["white"])
    c2.alignment = align(h="center")
    c2.border = thin_border()

ws.row_dimensions[6].height = 6  # spacer

# ── By module breakdown ────────────────────────────────────────────────────────
modules = {}
for r in results:
    m = r.get("area", "Unknown")
    modules.setdefault(m, {"PASS":0,"FAIL":0,"ERROR":0,"SKIP":0,"total":0})
    modules[m][r["status"]] = modules[m].get(r["status"], 0) + 1
    modules[m]["total"] += 1

row = 7
for col, lbl in enumerate(["Module", "Total", "PASS", "FAIL", "ERROR", "SKIP", "Pass %"], 1):
    hdr(ws, row, col, lbl, bg=CLR["blue"])
ws.row_dimensions[row].height = 22
row += 1

for i, (mod, counts) in enumerate(modules.items()):
    bg = CLR["grey"] if i % 2 == 0 else CLR["white"]
    t = counts["total"]
    p = counts["PASS"]
    pct = f"{p/t*100:.0f}%" if t else "-"
    cell(ws, row, 1, mod, bg=bg, bold=True)
    cell(ws, row, 2, t,   bg=bg, h="center")
    cell(ws, row, 3, p,   bg=bg, h="center", color=CLR["pass"] if p else "000000", bold=bool(p))
    cell(ws, row, 4, counts["FAIL"],  bg=bg, h="center", color=CLR["fail"] if counts["FAIL"] else "000000")
    cell(ws, row, 5, counts["ERROR"], bg=bg, h="center", color="C55A11" if counts["ERROR"] else "000000")
    cell(ws, row, 6, counts["SKIP"],  bg=bg, h="center")
    cell(ws, row, 7, pct, bg=bg, h="center", bold=True,
         color=CLR["pass"] if p == t else (CLR["fail"] if p == 0 else "000000"))
    ws.row_dimensions[row].height = 18
    row += 1

for col, w in enumerate([32, 8, 8, 8, 8, 8, 10], 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2 – ALL RESULTS
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Results")

headers = ["TC #", "Module / Area", "Scenario", "Status", "Expected", "Actual Result", "Notes", "Timestamp"]
widths  = [10, 24, 40, 14, 36, 48, 36, 20]

for i, (h, w) in enumerate(zip(headers, widths), 1):
    hdr(ws2, 1, i, h)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 26
ws2.freeze_panes = "A2"

for i, r in enumerate(results):
    row = i + 2
    bg = CLR["grey"] if i % 2 == 0 else CLR["white"]
    cell(ws2, row, 1, r.get("tc_id",""),     bg=bg, bold=True, h="center", color=CLR["blue"])
    cell(ws2, row, 2, r.get("area",""),      bg=bg, bold=True)
    cell(ws2, row, 3, r.get("scenario",""),  bg=bg, wrap=True)
    status_cell(ws2, row, 4, r.get("status",""))
    cell(ws2, row, 5, r.get("expected",""),  bg=bg, wrap=True)
    cell(ws2, row, 6, r.get("actual",""),    bg=bg, wrap=True)
    cell(ws2, row, 7, r.get("notes",""),     bg=bg, wrap=True,
         color=CLR["fail"] if r.get("status") in ("FAIL","ERROR") else "000000")
    ts = r.get("timestamp","")[:16].replace("T"," ")
    cell(ws2, row, 8, ts, bg=bg, h="center")
    ws2.row_dimensions[row].height = 52

# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3 – FAILURES ONLY
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Failures & Errors")

ws3.merge_cells("A1:G1")
c = ws3["A1"]
c.value = "FAILURES & ERRORS — Items Requiring Attention"
c.font = Font(name="Arial", bold=True, size=13, color=CLR["white"])
c.fill = fill(CLR["fail"])
c.alignment = align(h="center")
ws3.row_dimensions[1].height = 30

fail_headers = ["TC #", "Module", "Scenario", "Status", "Expected", "Actual / Root Cause", "Screenshot"]
fail_widths  = [10, 24, 40, 12, 36, 60, 40]
for i, (h, w) in enumerate(zip(fail_headers, fail_widths), 1):
    hdr(ws3, 2, i, h, bg=CLR["dgrey"])
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[2].height = 22
ws3.freeze_panes = "A3"

fail_results = [r for r in results if r["status"] in ("FAIL","ERROR")]
if not fail_results:
    ws3.merge_cells("A3:G3")
    c = ws3["A3"]
    c.value = "🎉  No failures or errors recorded!"
    c.font = Font(name="Arial", bold=True, size=12, color=CLR["pass"])
    c.alignment = align(h="center")
else:
    for i, r in enumerate(fail_results):
        row = i + 3
        bg = CLR["yellow"]
        cell(ws3, row, 1, r.get("tc_id",""),    bg=bg, bold=True, h="center", color=CLR["fail"])
        cell(ws3, row, 2, r.get("area",""),      bg=bg, bold=True)
        cell(ws3, row, 3, r.get("scenario",""),  bg=bg, wrap=True)
        status_cell(ws3, row, 4, r.get("status",""))
        cell(ws3, row, 5, r.get("expected",""),  bg=bg, wrap=True)
        cell(ws3, row, 6, r.get("actual",""),    bg=bg, wrap=True, color=CLR["fail"])
        cell(ws3, row, 7, r.get("screenshot",""), bg=bg, wrap=True, color="0563C1")
        ws3.row_dimensions[row].height = 60

out = "Job_Profit_OS_Test_Results.xlsx"
wb.save(out)
print(f"\n✅  Saved: {out}")
print(f"    {total} tests | {passed} PASS | {failed} FAIL | {errors} ERROR | {skipped} SKIP | {pass_rate} pass rate")
